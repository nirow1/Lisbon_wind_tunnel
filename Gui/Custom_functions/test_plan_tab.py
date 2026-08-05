import openpyxl
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import QWidget, QFileDialog, QTableWidgetItem, QTableWidget

from Qt_files.Qt_python.ui_test_plan_widget import Ui_Form


class TestPlanTab(QWidget):
    def __init__(self, columns: list):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self._init_ui()
        self.bind_buttons()
        self.add_columns(columns)

    def _init_ui(self):
        self.ui.xml_file_dir_btn.setIcon(QIcon("./App_data/dir_icon.png"))
        self.ui.test_running_wg.setVisible(False)
        self.ui.tableWidget.verticalHeader().setVisible(False)

    def bind_buttons(self):
        self.ui.xml_file_dir_btn.clicked.connect(self._open_file_dialog)
        self.ui.add_3d_row_btn.clicked.connect(self._add_row)
        self.ui.delete_3d_row_btn.clicked.connect(self._delete_row)

    def add_columns(self, columns: list):
        current_count = self.ui.tableWidget.columnCount()
        self.ui.tableWidget.setColumnCount(current_count + len(columns))
        for index, column in enumerate(columns, start=current_count):
            self.ui.tableWidget.setHorizontalHeaderItem(index, QTableWidgetItem(column))

    def get_test_plan(self) -> list:
        return self._create_test_plan(self.ui.tableWidget)

    @staticmethod
    def _create_test_plan(table: QTableWidget) -> list:
        test_plan = []
        for r in range(table.rowCount()):
            row_values = [
                table.item(r, c).text() if table.item(r, c) is not None else ""
                for c in range(table.columnCount())
            ]

            if not any(row_values):
                continue

            seconds = float(row_values[0]) if row_values[0] != "" else 0
            requested_values = tuple(float(v) if v != "" else "" for v in row_values[1:])
            test_plan.append((seconds, *requested_values))

        return test_plan

    def show_message(self, state: bool):
        self.ui.test_running_wg.setVisible(state)

    def _open_file_dialog(self):
        options = QFileDialog(self).options()
        self.xls_path = QFileDialog.getOpenFileName(self, "Select Folder", "", options=options)
        self.ui.xml_file_dir_le.setText(self.xls_path[0])
        if self.xls_path[0] != "":
            self._load_xls_data()

    def _load_xls_data(self):
        workbook = openpyxl.load_workbook(self.xls_path[0])
        sheet = workbook.worksheets[0]

        rows = sheet.max_row - 1
        self.ui.tableWidget.setRowCount(rows)

        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
            for j, cell in enumerate(row):
                cell = cell.value if cell.value is not None else ""
                self.ui.tableWidget.setItem(i, j, QTableWidgetItem(str(cell)))

        workbook.close()

    def _add_row(self):
        self.ui.tableWidget.insertRow(self.ui.tableWidget.rowCount())

    def _delete_row(self):
        row_count = self.ui.tableWidget.rowCount()
        if row_count > 0:
            self.ui.tableWidget.removeRow(row_count - 1)